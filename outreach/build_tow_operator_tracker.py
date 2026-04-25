"""
RIN — MTO Tow Operator Outreach Tracker Builder
Generates: RIN_Tow_Operator_Tracker.xlsx
Run: python3 build_tow_operator_tracker.py
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Colours ──────────────────────────────────────────────────────────────────
NAVY       = "05091A"
HEADER_BG  = "0C1428"
HEADER_FG  = "F0F4FF"

ZONE_A     = "1A3A2A"   # Cambridge / Kitchener — dark green
ZONE_B     = "1A2A3A"   # Milton / Oakville — dark blue
ZONE_C     = "2A1A3A"   # Brampton / Mississauga — dark purple
ZONE_D     = "3A2A1A"   # Stratford / SW Ontario — dark amber
ZONE_E     = "3A1A1A"   # Niagara / Beamsville — dark red
ZONE_F     = "1E1E1E"   # GTA general — dark grey

PROFILE_A  = "163A2A"   # Company owner — green tint
PROFILE_B  = "1A2838"   # Owner-operator — blue tint

STATUS_NEW       = "1E3A1E"   # New — dark green
STATUS_CALLED    = "1E2838"   # Called — dark blue
STATUS_QUALIFIED = "2E3A1E"   # Qualified — lime
STATUS_REGISTERED = "1E3820"  # Registered in DB — bright green
STATUS_DECLINED  = "3A1E1E"   # Declined — dark red
STATUS_VOICEMAIL = "2A2A1E"   # Left voicemail — yellow

WHITE = "F0F4FF"
GREY  = "9AABC8"

def fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def font(bold=False, color=WHITE, size=11):
    return Font(bold=bold, color=color, name="Inter", size=size)

def border():
    side = Side(style="thin", color="1E2D50")
    return Border(left=side, right=side, top=side, bottom=side)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

# ── Tow Operator Data ─────────────────────────────────────────────────────────
# Format: Zone, Profile, Company, Contact, Phone, MTO#, Truck Types,
#         Coverage Area, Hours, Motor Club, Avg Rate, Called Date,
#         Outcome, Registered, Notes

operators = [
    # Zone A — Cambridge / Kitchener (covers Diversco, Woodcock, Beyond, Frontier)
    ["Cambridge/KW", "A-Company", "Cambridge Towing & Recovery", "", "(519) 621-XXXX", "", "Flatbed, Wheel-lift", "Cambridge, Kitchener, Waterloo", "24/7", "CAA", "", "", "New", "No", "Find via google maps 'towing cambridge on'"],
    ["Cambridge/KW", "A-Company", "Waterloo Towing Service", "", "(519) 886-XXXX", "", "Flatbed, Heavy Recovery", "Waterloo, Kitchener, Guelph", "24/7", "CAA, Insurance", "", "", "New", "No", ""],
    ["Cambridge/KW", "B-Owner-Op", "(find via MTO registry)", "", "", "", "Flatbed", "Cambridge area", "Days", "", "", "", "New", "No", "Pull 2-3 from TSSEA registry for Cambridge"],
    ["Cambridge/KW", "B-Owner-Op", "(find via MTO registry)", "", "", "", "Flatbed", "Kitchener area", "Days", "", "", "", "New", "No", "Pull from TSSEA registry for Kitchener"],

    # Zone B — Milton / Oakville (covers ONE for Freight, Tandet Dedicated)
    ["Milton/Oakville", "A-Company", "Milton Towing", "", "(905) 876-XXXX", "", "Flatbed, Wheel-lift", "Milton, Halton Hills, Georgetown", "24/7", "CAA", "", "", "New", "No", ""],
    ["Milton/Oakville", "B-Owner-Op", "(find via MTO registry)", "", "", "", "Flatbed", "Oakville area", "Days", "", "", "", "New", "No", "Pull from TSSEA for Oakville/Milton"],

    # Zone C — Brampton / Mississauga (covers Trans4, BSD, Apps Transport)
    ["Brampton/Mississ.", "A-Company", "Brampton Towing Co.", "", "(905) 456-XXXX", "", "Flatbed, Wheel-lift, Heavy", "Brampton, Mississauga, Etobicoke", "24/7", "CAA, Insurance", "", "", "New", "No", ""],
    ["Brampton/Mississ.", "A-Company", "Peel Region Towing", "", "(905) 789-XXXX", "", "Flatbed, Wheel-lift", "Brampton, Caledon, Bolton", "24/7", "CAA", "", "", "New", "No", ""],
    ["Brampton/Mississ.", "B-Owner-Op", "(find via MTO registry)", "", "", "", "Flatbed", "Mississauga area", "Days", "", "", "", "New", "No", "Pull from TSSEA for Peel Region"],

    # Zone D — Stratford / SW Ontario (covers Steed Standard)
    ["Stratford/SW Ont.", "A-Company", "Stratford Towing & Storage", "", "(519) 271-XXXX", "", "Flatbed, Wheel-lift", "Stratford, St. Marys, Mitchell", "24/7", "CAA", "", "", "New", "No", ""],
    ["Stratford/SW Ont.", "B-Owner-Op", "(find via MTO registry)", "", "", "", "Flatbed", "Perth County", "Days", "", "", "", "New", "No", "Pull from TSSEA for Perth County"],

    # Zone E — Niagara / Beamsville (covers Spring Creek Carriers)
    ["Niagara/Beamsville", "A-Company", "Niagara Falls Towing", "", "(905) 356-XXXX", "", "Flatbed, Wheel-lift", "Niagara Falls, St. Catharines, Beamsville", "24/7", "CAA, Insurance", "", "", "New", "No", ""],
    ["Niagara/Beamsville", "B-Owner-Op", "(find via MTO registry)", "", "", "", "Flatbed", "Lincoln / Beamsville area", "Days", "", "", "", "New", "No", ""],

    # Zone F — Highway 401 / 400 Corridors (breakdown happens in transit)
    ["Hwy 401/400 corridor", "A-Company", "Highway 401 Towing", "", "", "", "Flatbed, Heavy Recovery", "Highway 401 GTA–Cambridge", "24/7", "OPP contract, Insurance", "", "", "New", "No", "OPP-contracted operators cover highways — high value"],
    ["Hwy 401/400 corridor", "A-Company", "(find via MTO registry)", "", "", "", "Heavy Recovery, Flatbed", "Highway 400 corridor", "24/7", "", "", "", "New", "No", "Look for operators near Barrie, Newmarket, Vaughan"],
]

COLUMNS = [
    ("Zone", 18),
    ("Profile", 14),
    ("Company Name", 28),
    ("Contact Name", 20),
    ("Phone", 18),
    ("MTO Cert #", 14),
    ("Truck Types", 24),
    ("Coverage Area", 26),
    ("Hours", 12),
    ("Motor Club", 18),
    ("Avg Flatbed Rate", 16),
    ("Called Date", 14),
    ("Outcome", 16),
    ("Registered in DB", 16),
    ("Notes", 40),
]

ZONE_FILLS = {
    "Cambridge/KW":       ZONE_A,
    "Milton/Oakville":    ZONE_B,
    "Brampton/Mississ.":  ZONE_C,
    "Stratford/SW Ont.":  ZONE_D,
    "Niagara/Beamsville": ZONE_E,
    "Hwy 401/400 corridor": ZONE_F,
}

OUTCOME_FILLS = {
    "New":        STATUS_NEW,
    "Called":     STATUS_CALLED,
    "Qualified":  STATUS_QUALIFIED,
    "Registered": STATUS_REGISTERED,
    "Declined":   STATUS_DECLINED,
    "Voicemail":  STATUS_VOICEMAIL,
}

def build():
    wb = openpyxl.Workbook()

    # ── Sheet 1: Operator Tracker ─────────────────────────────────────────────
    ws = wb.active
    ws.title = "Tow Operator Outreach"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "C4"

    # Title row
    ws.merge_cells("A1:O1")
    t = ws["A1"]
    t.value = "RIN — MTO Tow Operator Outreach Tracker · Q2 2026 Supply Pipeline"
    t.font = Font(bold=True, color=WHITE, name="Inter", size=13)
    t.fill = fill(NAVY)
    t.alignment = center()
    ws.row_dimensions[1].height = 32

    # Subtitle row
    ws.merge_cells("A2:O2")
    s = ws["A2"]
    s.value = "Goal: 8–10 registered operators across GTA zones before first fleet pilot dispatch"
    s.font = Font(bold=False, color=GREY, name="Inter", size=11)
    s.fill = fill(HEADER_BG)
    s.alignment = center()
    ws.row_dimensions[2].height = 22

    # Header row
    ws.row_dimensions[3].height = 28
    for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = font(bold=True, size=10)
        cell.fill = fill(HEADER_BG)
        cell.alignment = center()
        cell.border = border()
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Data rows
    for row_idx, op in enumerate(operators, start=4):
        zone      = op[0]
        profile   = op[1]
        outcome   = op[12]
        registered = op[13]

        zone_color    = ZONE_FILLS.get(zone, ZONE_F)
        outcome_color = OUTCOME_FILLS.get(outcome, STATUS_NEW)
        profile_color = PROFILE_A if profile.startswith("A") else PROFILE_B

        ws.row_dimensions[row_idx].height = 22

        for col_idx, value in enumerate(op, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = font(size=10)
            cell.border = border()

            # Column-specific styling
            if col_idx == 1:   # Zone
                cell.fill = fill(zone_color)
                cell.alignment = center()
            elif col_idx == 2:  # Profile
                cell.fill = fill(profile_color)
                cell.alignment = center()
            elif col_idx == 13: # Outcome
                cell.fill = fill(outcome_color)
                cell.alignment = center()
            elif col_idx == 14: # Registered
                cell.fill = fill(STATUS_REGISTERED if value == "Yes" else "2A1A1A")
                cell.alignment = center()
            elif col_idx == 15: # Notes
                cell.fill = fill(NAVY)
                cell.alignment = left()
            else:
                cell.fill = fill(NAVY)
                cell.alignment = left()

    # ── Sheet 2: Phone Script & Legend ────────────────────────────────────────
    ws2 = wb.create_sheet("Phone Script & Legend")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 70

    def s2row(row, label, value, label_bold=False):
        ws2.row_dimensions[row].height = 20
        a = ws2.cell(row=row, column=1, value=label)
        b = ws2.cell(row=row, column=2, value=value)
        a.font = Font(bold=label_bold, color=GREY if not label_bold else WHITE, name="Inter", size=10)
        b.font = Font(color=WHITE, name="Inter", size=10)
        a.fill = fill(HEADER_BG)
        b.fill = fill(NAVY)
        a.alignment = left()
        b.alignment = left()
        a.border = border()
        b.border = border()

    def s2title(row, text):
        ws2.row_dimensions[row].height = 28
        ws2.merge_cells(f"A{row}:B{row}")
        c = ws2.cell(row=row, column=1, value=text)
        c.font = Font(bold=True, color=WHITE, name="Inter", size=12)
        c.fill = fill(HEADER_BG)
        c.alignment = left()
        c.border = border()

    r = 1
    s2title(r, "PHONE SCRIPT — 30-Second Opener"); r+=1
    s2row(r, "Opening", "\"Hi, I'm Barnes with RIN — we're a dispatch platform connecting fleet trucks to tow operators in Ontario.\""); r+=1
    s2row(r, "The offer", "\"When a fleet truck breaks down in your zone, we send you an SMS job offer. You reply YES. Do the job. Get paid in 7 days. No monthly fee.\""); r+=1
    s2row(r, "The hook", "\"We're signing up operators in [their city] for a Q2 pilot. Do you handle commercial flatbed jobs?\""); r+=1
    s2row(r, "If YES", "\"Great. What's your typical rate for a standard flatbed tow in [city]?\" → Then: \"We'd keep 20% as a platform fee — you'd get 80% paid in 7 days guaranteed. Interested?\""); r+=1
    s2row(r, "If YES again", "\"Perfect. I'll send you a registration link by SMS right now — takes 3 minutes. What's the best number?\""); r+=1
    s2row(r, "If NO / busy", "\"No problem — I'll send you a quick text with the details. What's your mobile?\" → Send SMS with registration page link"); r+=1
    s2row(r, "If voicemail", "\"Hi, Barnes with RIN dispatch platform. We pay tow operators in 7 days for fleet jobs in your area. No monthly fee. Call or text [number] or visit [registration URL]. Thanks.\""); r+=1

    r+=1
    s2title(r, "QUALIFYING QUESTIONS (after opener)"); r+=1
    s2row(r, "Q1 — Truck types", "\"What types of trucks do you run? Flatbed, wheel-lift, heavy recovery?\""); r+=1
    s2row(r, "Q2 — Coverage zone", "\"What municipalities do you cover? Are you comfortable going to the 401?\""); r+=1
    s2row(r, "Q3 — Hours", "\"Do you operate 24/7 or days only?\""); r+=1
    s2row(r, "Q4 — Current work", "\"Are you affiliated with CAA or any insurance programs right now?\""); r+=1
    s2row(r, "Q5 — Rate check", "\"What's your standard flatbed rate in [city]?\" (validates our $260 avg)"); r+=1
    s2row(r, "Q6 — Capacity", "\"How many jobs a day could you take on from us realistically?\""); r+=1

    r+=1
    s2title(r, "PROFILE GUIDE"); r+=1
    s2row(r, "Profile A (Company)", "Owns 3–15 tow trucks. Has dispatchers/contracted drivers. One account, multiple trucks. Higher volume capacity.", label_bold=True); r+=1
    s2row(r, "Profile B (Owner-Op)", "Owns 1–3 trucks. IS the driver. Handles dispatch themselves. Lower volume but fast to onboard.", label_bold=True); r+=1

    r+=1
    s2title(r, "OUTCOME KEY"); r+=1
    s2row(r, "New", "Not yet contacted"); r+=1
    s2row(r, "Called", "Phone attempted, no answer or brief conversation"); r+=1
    s2row(r, "Voicemail", "Left voicemail — follow up in 48hrs"); r+=1
    s2row(r, "Qualified", "Interested, agreed to register — send link"); r+=1
    s2row(r, "Registered", "Submitted onboarding form — load into Supabase drivers table"); r+=1
    s2row(r, "Declined", "Not interested or wrong fit — note reason"); r+=1

    r+=1
    s2title(r, "ZONE COVERAGE MAP"); r+=1
    s2row(r, "Cambridge/KW", "Covers: Diversco, Woodcock Brothers, Beyond Transportation, Frontier Distribution"); r+=1
    s2row(r, "Milton/Oakville", "Covers: ONE for Freight, Tandet Dedicated (Oakville)"); r+=1
    s2row(r, "Brampton/Mississ.", "Covers: Trans4 Group, BSD Linehaul, Apps Transport, ASL Global"); r+=1
    s2row(r, "Stratford/SW Ont.", "Covers: Steed Standard Transport"); r+=1
    s2row(r, "Niagara/Beamsville", "Covers: Spring Creek Carriers"); r+=1
    s2row(r, "Hwy 401/400 corridor", "Covers: In-transit breakdowns — highest frequency zone"); r+=1

    r+=1
    s2title(r, "SUPABASE FIELDS TO FILL (drivers table)"); r+=1
    s2row(r, "company_name", "Towing company name"); r+=1
    s2row(r, "driver_name", "Primary contact name"); r+=1
    s2row(r, "phone", "SMS-capable dispatch phone — CRITICAL for dispatch"); r+=1
    s2row(r, "service_radius_km", "Based on coverage zone (typically 30–80km)"); r+=1
    s2row(r, "availability_status", "Set to 'available' on registration"); r+=1
    s2row(r, "reliability_score", "Start at 3.0 (default) — updates after first jobs"); r+=1
    s2row(r, "rating", "Start at 0.0 — updates after first customer ratings"); r+=1

    wb.save("/Users/b.lamoutlook.com/Downloads/rin-marketing/outreach/RIN_Tow_Operator_Tracker.xlsx")
    print("✓ RIN_Tow_Operator_Tracker.xlsx saved")

if __name__ == "__main__":
    build()
