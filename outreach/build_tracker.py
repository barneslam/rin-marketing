import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "RIN Fleet Outreach"

# ── Colours ──────────────────────────────────────────────────────────────────
DARK_NAVY   = "0D1B2A"
MID_BLUE    = "1E3A5F"
ACCENT_BLUE = "2563EB"
LIGHT_BLUE  = "DBEAFE"
TIER1_BG    = "EFF6FF"
TIER2_BG    = "F0FDF4"
TIER3_BG    = "FFF7ED"
WARN_BG     = "FEF2F2"
WARN_FG     = "991B1B"
CONF_GREEN  = "166534"
CONF_BG     = "DCFCE7"
INF_BG      = "FEF9C3"
INF_FG      = "854D0E"
GEN_BG      = "F3F4F6"
GEN_FG      = "374151"
WHITE       = "FFFFFF"
HEADER_FG   = WHITE

def cell_style(fill_hex, font_hex=None, bold=False, size=10, wrap=False, halign="left", valign="top"):
    fill = PatternFill("solid", fgColor=fill_hex)
    font = Font(color=font_hex or "000000", bold=bold, size=size, name="Calibri")
    align = Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)
    return fill, font, align

thin = Side(style="thin", color="CBD5E1")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ── Title row ─────────────────────────────────────────────────────────────────
ws.merge_cells("A1:O1")
title_cell = ws["A1"]
title_cell.value = "RIN Fleet Outreach Tracker — Q2 2026"
title_cell.font = Font(name="Calibri", bold=True, size=16, color=WHITE)
title_cell.fill = PatternFill("solid", fgColor=DARK_NAVY)
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 36

# ── Sub-header row ────────────────────────────────────────────────────────────
ws.merge_cells("A2:O2")
sub = ws["A2"]
sub.value = "Max 5 emails/day  •  No Calendly on first touch  •  Follow-up: 5–7 days after send  •  One follow-up max  •  Batch 3: hold until 2+ pilots live"
sub.font = Font(name="Calibri", italic=True, size=9, color="94A3B8")
sub.fill = PatternFill("solid", fgColor=MID_BLUE)
sub.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 18

# ── Column headers ────────────────────────────────────────────────────────────
headers = [
    "Batch", "Company", "Fleet Size", "Contact Name", "Title",
    "Email", "Email Confidence", "Phone", "Website",
    "Pain Point", "Template", "Date Sent", "Follow-up Date",
    "Status", "Notes"
]
col_widths = [8, 30, 14, 22, 26, 34, 14, 16, 26, 48, 12, 14, 14, 16, 40]

for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
    cell = ws.cell(row=3, column=col_idx, value=h)
    cell.fill = PatternFill("solid", fgColor=ACCENT_BLUE)
    cell.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
    ws.column_dimensions[get_column_letter(col_idx)].width = w
ws.row_dimensions[3].height = 30

# ── Data ──────────────────────────────────────────────────────────────────────
# Columns: batch, company, fleet_size, contact_name, title, email,
#          email_conf, phone, website, pain_point, template,
#          date_sent, followup_date, status, notes
rows = [
    # ── BATCH 1 (Tier 2 — Mid-size, Send First) ──
    ("Batch 1", "Trans4 Group Ltd.",
     "~16 trucks", "Simrat Singh", "Manager, Operations",
     "s.singh@trans4.ca", "Inferred", "(855) 633-5444", "trans4.ca",
     "At 16 trucks, one vehicle off the road is a real percentage of capacity. The ops manager personally manages every breakdown with no dispatch buffer.",
     "ops_manager", "", "", "Identified",
     "Brampton ON. CEO Robert McDougall also at r.mcdougall@trans4.ca. Small fleet — owner-level pain."),

    ("Batch 1", "ONE for Freight",
     "~25 trucks", "Luke Robak", "General Manager",
     "l.robak@oneforfreight.com", "Inferred", "(905) 876-3996", "oneforfreight.com",
     "Cross-border LTL to New England means a breakdown in the wrong spot ripples into customs timelines and receiver commitments — not just a tow call.",
     "ops_manager", "", "", "Identified",
     "Milton ON. Toll free 1-800-363-5143. 8450 Lawson Rd Unit 1&2."),

    ("Batch 1", "Diversco Transport Inc.",
     "~30 trucks", "Janet Huber", "Operations/Safety Manager",
     "jhuber@diversco.com", "CONFIRMED", "(226) 972-8708", "diversco.ca",
     "Janet owns both ops and safety — a breakdown drops into her lap with no backup. Every hour managing a tow is an hour not running the operation.",
     "ops_manager", "", "", "Identified",
     "Cambridge ON. dispatch@diversco.com also works. Toll free (800) 387-8544. BEST FIRST SEND — email confirmed."),

    ("Batch 1", "J&R Hall Transport Inc.",
     "~105 trucks", "— (no public contact)", "Call first → ask for Ops/Fleet Mgr",
     "info@jrhall.ca", "General", "(519) 632-7429", "jrhall.ca",
     "4th generation family carrier running expedited LTL and FTL. 40% team-driven. Breakdown on an expedited load has a direct customer penalty.",
     "ops_manager", "", "", "Identified",
     "552 Piper St Ayr ON N0B 1E0. Call first to get a named contact before emailing."),

    ("Batch 1", "B.S.D. Linehaul",
     "~89-101 trucks", "Sukhwinder Chandi", "Vice President",
     "dispatch@bsdlinehaul.com", "CONFIRMED", "(905) 499-2811", "bsdlinehaul.com",
     "Heavy-haul and tandem fleet means bigger vehicles with more complex roadside needs — and higher cost when they go down.",
     "dispatcher", "", "", "Identified",
     "Brampton ON. Also Randall Tuck / Rajwinder Kaur (Dispatch Coordinators) at same domain."),

    ("Batch 1", "Woodcock Brothers Transportation",
     "100+ units", "Gerald Bice", "Manager, Dispatch",
     "g.bice@woodcockbrothers.com", "Inferred", "1-800-565-5557", "woodcockbrothers.com",
     "Mixed van / flatbed / step-deck fleet across Ontario means variable breakdown scenarios. Gerald owns dispatch — he's managing the call every time.",
     "dispatcher", "", "", "Identified",
     "225 Huron Rd Sebringville ON N0K 1X0. Rebranded Woodcock Transportation Group 2020. President: Brad Woodcock."),

    ("Batch 1", "ASL Global Logistics",
     "Unknown", "— (no public contact)", "Call first → ask for VP Operations",
     "info@aslgloballogistics.com", "General", "(416) 243-3010", "aslgloballogistics.ca",
     "GTA-based logistics operator — unknown fleet size but 150K sq ft warehouse suggests steady freight volume and regular road exposure.",
     "finance", "", "", "Identified",
     "300 Kennedy Rd S Unit D Brampton ON L6T 2H4. Call and ask for VP Operations before emailing."),

    ("Batch 1", "Onfreight Logistics",
     "~85 trucks", "Ramandeep Kaur", "Dispatch Supervisor",
     "r.kaur@onfreight.ca", "Inferred", "(519) 727-4578", "onfreight.ca",
     "Windsor-area carrier — the 401 corridor is breakdown territory. Ramandeep owns dispatch and personally absorbs every roadside event.",
     "dispatcher", "", "", "Identified",
     "302 Patillo Rd Tecumseh ON N8N 2L9. Also Jim Pereira Fleet Safety Mgr j.pereira@onfreight.ca. President Steve Ondejko is OTA Chairman."),

    # ── BATCH 2 (Tier 1 — Enterprise, Send Second) ──
    ("Batch 2", "Canada Cartage Systems Ltd.",
     "~4,000 units", "Roger Matchett", "Director, Fleet Operations",
     "roger.matchett@canadacartage.com", "Inferred", "(416) 740-3000", "canadacartage.com",
     "Largest dedicated fleet operator in Canada — 4,000+ units means dozens of roadside events per week. Invoice predictability and audit trail is the real sell.",
     "ops_manager", "", "", "Identified",
     "Also absorbed Walmart Fleet ULC (Jan 2025). Finance angle: Salina Dias Director Operations. Massive org — long sales cycle expected."),

    ("Batch 2", "JD Smith Supply Chain Solutions",
     "Est. 50-150 trucks", "John Kupiec", "VP Transportation",
     "john.kupiec@jdsmith.com", "Inferred", "(905) 669-8980", "jdsmith.com",
     "GTA-based 3PL since 1919 — fleet serves pharma, food, CPG. Breakdown in a temperature-controlled run has spoilage liability on top of downtime cost.",
     "ops_manager", "", "", "Identified",
     "Vaughan ON. Email format firstname.lastname@jdsmith.com confirmed. Alt: jdsmith.com/contact-us web form."),

    ("Batch 2", "Fortigo Freight Services Inc.",
     "Est. 50-100 trucks", "Alex Koretskyi", "Fleet Manager",
     "a.koretskyi@fortigofreight.com", "Inferred", "(416) 367-8446", "fortigofreight.com",
     "Best Fleets to Drive For winner — high ops standards mean they care about service quality. Alex owns fleet directly. Finance: $19.5M revenue.",
     "dispatcher", "", "", "Identified",
     "50 Belfield Rd Etobicoke ON M9W 1G1. Also Christoff Pierre Logistics Mgr c.pierre@fortigofreight.com."),

    ("Batch 2", "GX Transportation Solutions Inc.",
     "Est. 30-75 trucks", "D.J. Massicotte", "VP Operations",
     "info@gxts.com", "General", "(905) 455-4341", "gxts.com",
     "Multi-modal (road/rail/air) with structured ops team. VP Operations holds the budget — cost predictability and clean reporting is the angle.",
     "finance", "", "", "Identified",
     "6495 Van Deemter Ct Mississauga ON L5T 1S1. No direct email found — use info@gxts.com and address DJ Massicotte by name."),

    ("Batch 2", "Tandet",
     "~169 trucks (2 divisions)", "Al Kruse", "Fleet Manager",
     "a.kruse@tandet.com", "Inferred", "1-888-826-3388", "tandet.com",
     "Two Ontario divisions (Sarnia + Oakville) mean split ops — Al's role spans fleet across locations. Coordination overhead when a truck goes down off-site.",
     "ops_manager", "", "", "Identified",
     "1925A Barton St E Hamilton ON L8H 2Y7. Also Doug Bowes + John Wiles (Ops Mgrs). General: info@tandet.com."),

    ("Batch 2", "XTL Transport Inc.",
     "~500 trucks + 1,300 trailers", "Darcy MacArthur", "VP Logistics Operations",
     "darcy.macarthur@xtl.com", "CONFIRMED", "(416) 742-2345 x9041", "xtl.com",
     "Large family-owned carrier with temp-controlled and specialized trailers — breakdowns on sensitive loads carry extra urgency. Darcy owns logistics ops.",
     "ops_manager", "", "", "Identified",
     "Etobicoke ON. Toll free 800-665-9318. Also Jason Fisher VP Transport Ops jason.fisher@xtl.com."),

    ("Batch 2", "Polaris Transportation Group",
     "~60+ trucks + 150 trailers", "Ira Basil", "Director, Safety Risk & Fleet",
     "i.basil@polaristransport.com", "Inferred", "(905) 671-3100", "polaristransport.com",
     "Cross-border LTL specialists — US-Canada runs can't afford breakdown delays at border. Ira owns fleet + safety, which means every incident is his.",
     "dispatcher", "", "", "Identified",
     "7099 Torbram Rd Mississauga ON L4T 1G7. Acquired Drapeau (chemical haulers) May 2023 — fleet may be larger now."),

    ("Batch 2", "Cardinal Couriers",
     "135+ trucks", "— (no named contact)", "Call dispatch → ask for Ops/Dispatch Mgr",
     "salesinfo@cardinalcouriers.com", "General", "(905) 507-8844", "cardinalcouriers.com",
     "Pre-8am delivery window means zero margin for roadside delays. If a truck goes down at 5am, the whole delivery window fails.",
     "dispatcher", "", "", "Identified",
     "400 Brunel Rd Mississauga ON. Call dispatch line to get a named contact before emailing."),

    ("Batch 2", "Liberty Linehaul (via Kriska)",
     "~72 tractors", "Chris Gerber", "General Manager (Kriska/Liberty)",
     "c.gerber@kriskagroup.com", "Inferred", "(519) 740-7072", "kriskagroup.com",
     "Cross-border Ontario-California corridor — a breakdown mid-haul is a multi-day event. Acquired by Kriska 2023, still operates independently in Ayr.",
     "ops_manager", "", "", "Identified",
     "⚠️ ACQUIRED by Kriska (Mar 2023). Liberty Linehaul name retained. Chris Gerber is GM in Ayr ON."),

    ("Batch 2", "Loblaw Transport",
     "130-160+ trucks", "Paul James", "Director, National Transport Maintenance",
     "paul.james@loblaw.ca", "Inferred", "(905) 821-2111", "loblaw.ca",
     "Massive corporate fleet with active EV expansion. Breakdown cost visibility and clean monthly invoicing is the angle — not speed of dispatch.",
     "finance", "", "", "Identified",
     "1940 Argentia Rd Mississauga ON. Also Wayne Scott Sr Director Transport Maintenance w.scott@loblaw.ca. Long procurement cycle — expect gatekeeper."),

    # ── BATCH 3 (Watch List — Hold until 2+ pilots live) ──
    ("Batch 3 ⏸", "Autobahn Freight Lines Ltd.",
     "229-650+ trucks", "Prashob Peethambaran", "Manager, Fleet Compliance",
     "p.peethambaran@autobahnfreight.com", "Inferred", "(416) 741-5454", "autobahnfreight.com",
     "⚠️ Actually enterprise-scale (700+ drivers). Reconsider tier — may need enterprise pitch and longer cycle.",
     "ops_manager", "", "", "Hold",
     "300 Kennedy S Unit D Brampton ON L6W 4V2. Also info@autobahnfreight.com. Move to Batch 2 when ready."),

    ("Batch 3 ⏸", "Beyond Transportation Inc.",
     "~32 trucks", "— (no public contact)", "Call → ask for Operations Manager",
     "info@beyondtransportationinc.com", "General", "(519) 632-7781", "beyondtransportationinc.com",
     "Small Ontario carrier in Ayr — ops manager personally handles breakdowns. Good mid-size fit once pilot proof exists.",
     "ops_manager", "", "", "Hold",
     "300 Melair Drive Ayr ON N0B 1E0. Call to get named contact."),

    ("Batch 3 ⏸", "Frontier Distribution Services Inc.",
     "~25 trucks", "— (no public contact)", "Call → ask for Fleet Manager",
     "info@frontierdistribution.ca", "General", "(519) 696-2113", "frontierdistribution.ca",
     "Small Kitchener-based truckload carrier. Good fit — needs pilot proof first.",
     "ops_manager", "", "", "Hold",
     "38 Mcbrine Place Unit 3 Kitchener ON N2R 1G8."),

    ("Batch 3 ⏸", "Mill Creek Motor Freight LP",
     "~180 trucks", "Matthew Kaczmarski", "Director of Operations",
     "m.kaczmarski@millcreek.on.ca", "Inferred", "", "millcreek.on.ca",
     "Part of Kriska Transportation Group — operates independently out of Ayr ON. 180 trucks means steady incident volume.",
     "ops_manager", "", "", "Hold",
     "101 Earl Thompson Rd Ayr ON N0B 1E0. Also GM Lorie Thompson (Automotive Ops). Contact millcreek.on.ca/contact/"),

    ("Batch 3 ⏸", "Spring Creek Carriers Inc.",
     "~37 trucks", "Mark Bylsma", "President",
     "mark@springcreekcarriers.com", "CONFIRMED", "(905) 563-9989 x222", "springcreekcarriers.com",
     "Owner-operated carrier in Niagara — Mark personally manages every breakdown. Confirmed email. Good fit once pilot exists.",
     "ops_manager", "", "", "Hold",
     "4695 Bartlett Rd Beamsville ON L0R 1B1. Small owner-operated — email confirmed, go direct when ready."),

    ("Batch 3 ⏸", "Rosedale Transport",
     "500+ tractors + 1,300 trailers", "— (no named contact)", "Send to hello@ and request routing",
     "hello@rosedale.ca", "General", "(905) 670-0057", "rosedalegroup.com",
     "⚠️ Actually enterprise-scale — 15 terminals, 800+ employees. Needs enterprise pitch, not mid-size story.",
     "ops_manager", "", "", "Hold",
     "Mississauga ON HQ. Toll free 1-877-588-0057. Reclassify to Batch 2 enterprise when ready."),

    ("Batch 3 ⏸", "Steed Standard Transport Limited",
     "34-42 trucks", "James Steed", "President",
     "james.steed@sst.ca", "CONFIRMED", "(519) 271-9924", "sst.ca",
     "110-year-old carrier in Stratford — owner-operated, James personally manages the operation. Confirmed email, easy first touch when ready.",
     "ops_manager", "", "", "Hold",
     "603 Romeo St S Stratford ON N5A 6S5. 34-42 trucks + 100 trailers. Go direct to James."),

    ("Batch 3 ⏸", "Joseph Haulage Canada Corp.",
     "~200 units", "Geoffrey Joseph", "President & CEO",
     "g.joseph@josephhaulage.com", "Inferred", "", "josephhaulage.com",
     "Heavy-haul / industrial — dump trailers, flatbeds, tankers across 3 locations. Specialized equipment means complex roadside scenarios.",
     "ops_manager", "", "", "Hold",
     "590 S Service Rd Stoney Creek ON L8E 2W1. Call to find ops manager below Geoffrey."),

    # ── SKIP ──
    ("SKIP ⛔", "Walmart Fleet ULC",
     "N/A — absorbed", "Roger Matchett (now at Canada Cartage)", "Director Fleet Operations",
     "roger.matchett@canadacartage.com", "N/A", "(416) 740-3000", "canadacartage.com",
     "ACQUIRED by Canada Cartage (Jan 2025). Fleet fully absorbed. Do not send — covered by Canada Cartage email.",
     "—", "", "", "Skip",
     "Contact Canada Cartage once only. Do not duplicate."),
]

# Status colour map
STATUS_COLORS = {
    "Identified": ("EFF6FF", "1D4ED8"),
    "Contacted":  ("FEF9C3", "854D0E"),
    "Replied":    ("DCFCE7", "166534"),
    "Qualified":  ("F3E8FF", "6B21A8"),
    "Pilot":      ("ECFDF5", "065F46"),
    "Hold":       ("FFF7ED", "9A3412"),
    "Skip":       ("F9FAFB", "6B7280"),
}

CONF_COLORS = {
    "CONFIRMED": (CONF_BG, CONF_GREEN),
    "Inferred":  (INF_BG, INF_FG),
    "General":   (GEN_BG, GEN_FG),
    "N/A":       (GEN_BG, GEN_FG),
}

BATCH_COLORS = {
    "Batch 1": ("D1FAE5", "065F46"),
    "Batch 2": ("DBEAFE", "1E40AF"),
    "Batch 3 ⏸": ("FEF3C7", "92400E"),
    "SKIP ⛔": ("F3F4F6", "6B7280"),
}

for row_idx, row_data in enumerate(rows, start=4):
    batch = row_data[0]
    batch_bg, batch_fg = BATCH_COLORS.get(batch, ("FFFFFF", "000000"))

    # Row background
    row_bg = "FFFFFF" if row_idx % 2 == 0 else "F8FAFC"

    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = border
        cell.font = Font(name="Calibri", size=9)
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.fill = PatternFill("solid", fgColor=row_bg)

        # Batch column
        if col_idx == 1:
            cell.fill = PatternFill("solid", fgColor=batch_bg)
            cell.font = Font(name="Calibri", size=9, bold=True, color=batch_fg)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Email confidence column (col 7)
        if col_idx == 7:
            conf_bg, conf_fg = CONF_COLORS.get(value, (GEN_BG, GEN_FG))
            cell.fill = PatternFill("solid", fgColor=conf_bg)
            cell.font = Font(name="Calibri", size=9, bold=True, color=conf_fg)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Status column (col 14)
        if col_idx == 14:
            st_bg, st_fg = STATUS_COLORS.get(value, ("FFFFFF", "000000"))
            cell.fill = PatternFill("solid", fgColor=st_bg)
            cell.font = Font(name="Calibri", size=9, bold=True, color=st_fg)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Date columns — light yellow placeholder
        if col_idx in (12, 13):
            cell.fill = PatternFill("solid", fgColor="FFFBEB")
            cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[row_idx].height = 52

# ── Freeze panes at row 4, col 3 ─────────────────────────────────────────────
ws.freeze_panes = "C4"

# ── Auto-filter ───────────────────────────────────────────────────────────────
ws.auto_filter.ref = f"A3:O{3 + len(rows)}"

# ── Legend sheet ──────────────────────────────────────────────────────────────
ls = wb.create_sheet("Legend & Rules")
ls.column_dimensions["A"].width = 22
ls.column_dimensions["B"].width = 60

legend_header = ls["A1"]
legend_header.value = "RIN Outreach — Rules & Legend"
legend_header.font = Font(name="Calibri", bold=True, size=13, color=WHITE)
legend_header.fill = PatternFill("solid", fgColor=DARK_NAVY)
ls.merge_cells("A1:B1")
ls["A1"].alignment = Alignment(horizontal="center", vertical="center")
ls.row_dimensions[1].height = 28

rules = [
    ("SEND RULES", ""),
    ("Max per day", "5 emails/day to protect domain deliverability"),
    ("First touch", "No Calendly link on first send — add only after a positive reply"),
    ("Follow-up timing", "5–7 days after first send (auto-calc from Date Sent)"),
    ("Follow-up max", "One follow-up only. If no reply to follow-up, mark Dead."),
    ("Batch 3", "Hold until 2+ pilot conversations are live (need social proof)"),
    ("", ""),
    ("EMAIL CONFIDENCE", ""),
    ("CONFIRMED", "Email verified from public source or company page"),
    ("Inferred", "Name found, email pattern deduced from domain. Expect ~15-20% bounce rate."),
    ("General", "Inbox only (info@/dispatch@). Call first to get a named contact."),
    ("", ""),
    ("PIPELINE STAGES", ""),
    ("Identified", "In list, not yet contacted"),
    ("Contacted", "First email sent — enter date"),
    ("Replied", "Any reply received (positive or negative)"),
    ("Qualified", "15-min call booked"),
    ("Pilot", "Pilot terms sent or signed"),
    ("Hold", "Batch 3 — do not send yet"),
    ("Skip", "Acquired/duplicate — do not send"),
    ("", ""),
    ("BATCH PRIORITY", ""),
    ("Batch 1", "Mid-size (Tier 2) — send first. Personal ops pain. 8 active targets."),
    ("Batch 2", "Enterprise (Tier 1) — send after Batch 1. Finance/audit angle. 10 active targets."),
    ("Batch 3", "Watch list — hold until social proof exists. 8 targets."),
    ("", ""),
    ("CALENDLY LINKS", ""),
    ("15-min qualification", "calendly.com/barnes-lam/free-consultation-24-hour-business-sprint"),
    ("30-min deep dive", "calendly.com/barnes-lam/30-min-free-strategy-call-clone"),
    ("", ""),
    ("KEY FLAGS", ""),
    ("⚠️ Apps Transport", "Acquired by TFI International — confirm contact still valid before sending"),
    ("⚠️ Liberty Linehaul", "Acquired by Kriska (Mar 2023) — contact Chris Gerber via kriskagroup.com"),
    ("⛔ Walmart Fleet ULC", "Acquired by Canada Cartage (Jan 2025) — SKIP, covered by CC row"),
    ("⚠️ Autobahn Freight", "Actually enterprise-scale (229-650 trucks) — consider moving to Batch 2"),
    ("⚠️ Rosedale Transport", "Actually enterprise-scale (500+ trucks) — consider moving to Batch 2"),
]

for r, (key, val) in enumerate(rules, start=2):
    ka = ls.cell(row=r, column=1, value=key)
    va = ls.cell(row=r, column=2, value=val)
    if val == "":
        ka.fill = PatternFill("solid", fgColor=MID_BLUE)
        ka.font = Font(name="Calibri", bold=True, size=9, color=WHITE)
        ls.merge_cells(f"A{r}:B{r}")
        ls.row_dimensions[r].height = 18
    else:
        ka.font = Font(name="Calibri", bold=True, size=9)
        va.font = Font(name="Calibri", size=9)
        ka.alignment = Alignment(vertical="top", wrap_text=True)
        va.alignment = Alignment(vertical="top", wrap_text=True)
        ls.row_dimensions[r].height = 16

wb.save("/Users/b.lamoutlook.com/Downloads/rin-marketing/outreach/RIN_Fleet_Outreach_Tracker.xlsx")
print("Excel file created successfully.")
